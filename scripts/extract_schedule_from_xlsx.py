import json
import re
import unicodedata
from pathlib import Path

import openpyxl


def norm(value: str) -> str:
    value = unicodedata.normalize("NFD", value or "")
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def normalize_day(raw_day: str | None) -> str | None:
    n = norm(raw_day or "")
    if n.startswith("segundafeira"):
        return "segunda"
    if n.startswith("ter") and "feira" in n:
        return "terca"
    if n.startswith("quartafeira"):
        return "quarta"
    if n.startswith("quintafeira"):
        return "quinta"
    if n.startswith("sextafeira"):
        return "sexta"
    if n.startswith("sabado"):
        return "sabado"
    if n.startswith("domingo"):
        return "domingo"
    return None


DAY_LABEL = {
    "segunda": "Segunda-feira",
    "terca": "Terça-feira",
    "quarta": "Quarta-feira",
    "quinta": "Quinta-feira",
    "sexta": "Sexta-feira",
    "sabado": "Sábado",
    "domingo": "Domingo",
}
DAY_ORDER = ["segunda", "terca", "quarta", "quinta", "sexta", "sabado", "domingo"]


def main() -> None:
    base = Path.cwd()
    desktop = Path.home() / "Desktop"
    matches = list(desktop.glob("Hor*2026*.xlsx"))
    if not matches:
        raise RuntimeError("Planilha Hor*2026*.xlsx não encontrada na área de trabalho.")

    xlsx_path = matches[0]
    output_path = base / "tmp_schedule_all_from_planilha.json"

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Planilha1"]

    merged_lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                merged_lookup[(rr, cc)] = (min_row, min_col)

    def cell_text(row: int, col: int) -> str | None:
        mr, mc = merged_lookup.get((row, col), (row, col))
        value = ws.cell(mr, mc).value
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    # Header line with classes is line 4 in current spreadsheet.
    header_row = 4
    time_cols = {2, 9, 16}
    class_cols: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        if col in time_cols:
            continue
        label = cell_text(header_row, col)
        if not label:
            continue
        if norm(label) == "horario":
            continue
        class_cols[col] = label

    extracted: list[dict] = []
    current_day: str | None = None
    for row in range(5, ws.max_row + 1):
        day_cell = cell_text(row, 1)
        normalized_day = normalize_day(day_cell)
        if normalized_day:
            current_day = normalized_day
        if not current_day:
            continue

        block_slot: dict[int, str] = {}
        for tcol in (2, 9, 16):
            slot = cell_text(row, tcol)
            if slot:
                block_slot[tcol] = slot

        for col, class_label in class_cols.items():
            tcol = 2 if col <= 8 else (9 if col <= 15 else 16)
            slot = block_slot.get(tcol)
            if not slot:
                continue

            raw_content = cell_text(row, col)
            if not raw_content or raw_content == "*":
                continue

            time_match = re.match(r"^(\d{2}:\d{2})h?/(\d{2}:\d{2})h?$", slot)
            if not time_match:
                continue

            subject_raw = raw_content
            teacher_hint = None
            match_teacher = re.match(r"^(.*)\(([^()]*)\)\s*$", raw_content)
            if match_teacher:
                subject_raw = match_teacher.group(1).strip()
                teacher_hint = match_teacher.group(2).strip()

            extracted.append(
                {
                    "linha_planilha": row,
                    "class_label": class_label,
                    "dia_semana": DAY_LABEL[current_day],
                    "dia_ordem": DAY_ORDER.index(current_day) + 1,
                    "horario_raw": slot,
                    "hora_inicio": time_match.group(1),
                    "hora_fim": time_match.group(2),
                    "conteudo_raw": raw_content,
                    "disciplina_raw": subject_raw,
                    "professor_hint": teacher_hint,
                    "tipo": "INTERVALO" if norm(raw_content) == "intervalo" else "AULA",
                }
            )

    extracted.sort(
        key=lambda r: (norm(r["class_label"]), int(r["dia_ordem"]), str(r["hora_inicio"]))
    )

    payload = {
        "fonte": str(xlsx_path),
        "total_registros": len(extracted),
        "classes_detectadas": sorted({r["class_label"] for r in extracted}),
        "registros": extracted,
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_path)
    print(len(extracted))


if __name__ == "__main__":
    main()
